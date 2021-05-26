import pandas as pd
import numpy as np
import time
from time import time, sleep
import os
import xlsxwriter
import requests
from openpyxl import load_workbook
from datetime import datetime
import json



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,truncate_sheet=False, **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def choose_url(choice):
    if choice=='NIFTY' or choice=='BANKNIFTY':
        url = f'https://www.nseindia.com/api/option-chain-indices?symbol={choice}'

    else:
        url = f'https://www.nseindia.com/api/option-chain-equities?symbol={choice}'
    
    return url

def init_spreadsheet(choice,date,selected_margin):


    # headers = {"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.1.1 Safari/605.1.15","Accept-Language": "en-gb","Accept-Encoding":"br, gzip, deflate","Accept":"test/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Referer":"http://www.google.com/"}
    headers = {
        'Connection': 'keep-alive',
        'Cache-Control': 'max-age=0',
        'DNT': '1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
        'Sec-Fetch-User': '?1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-Mode': 'navigate',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
        }  

    for i in range(0,40):
        # headers = {"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.1.1 Safari/605.1.15","Accept-Language": "en-gb","Accept-Encoding":"br, gzip, deflate","Accept":"test/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Referer":"http://www.google.com/"}
        # url = "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY"
        try:
            page = requests.get(choose_url(choice),headers=headers).json()
            # print(output)
        except ValueError:
            s = requests.Session()
            page = s.get("http://nseindia.com",headers=headers)
            page = s.get(choose_url(choice),headers=headers).json()

        if len(page)>0:
            break

    res = page['records']['data'] #########
    underlying_value = int(res[0]['PE']['underlyingValue']) #######

    if choice not in ['NIFTY','BANKNIFTY']:
        selected_margin = underlying_value * selected_margin / 100

    data = [d.get('PE') for d in res]
    data_cleaned = list(filter(None, data))
    df_put = pd.DataFrame(data_cleaned)
    df_put['expiryDate'] = pd.to_datetime(df_put['expiryDate'])

    df_put.sort_values(by=['expiryDate','strikePrice'],inplace=True)

    col_list = ['strikePrice','expiryDate','openInterest']

    df_put = df_put[col_list]

    df_put = df_put[df_put['expiryDate'] == date]

    df_put['strikePrice'] = df_put['strikePrice'].map(int)

    df_put = df_put[(df_put['strikePrice'] >= underlying_value-selected_margin) & (df_put['strikePrice']<=underlying_value+selected_margin)]

    sp_range = df_put.strikePrice.unique().tolist()  

    data_format = ['Time','Stock/Index','Expiry','Right','Current OI','Previous Minute OI','OI Change',
                '% Up/Down from Last Minute OI','Right','Current OI','Previous Minute OI','OI Change',
                '% Up/Down from Last Minute OI']

    base_df = pd.DataFrame(columns = data_format)

    file_name = choice + ' ' + date
    for i in sp_range:
        append_df_to_excel(f"{file_name}.xlsx",base_df,sheet_name=str(i), index=False, truncate_sheet = True)
    
    return sp_range


def main(choice,previous_min_OI_x,previous_min_OI_y,selected_margin):

    headers = {
        'Connection': 'keep-alive',
        'Cache-Control': 'max-age=0',
        'DNT': '1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
        'Sec-Fetch-User': '?1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-Mode': 'navigate',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
        }   

    try:
    
        for i in range(0,40):
            # headers = {"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.1.1 Safari/605.1.15","Accept-Language": "en-gb","Accept-Encoding":"br, gzip, deflate","Accept":"test/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Referer":"http://www.google.com/"}
            # url = "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY"
            try:
                page = requests.get(choose_url(choice),headers=headers).json()
                # print(output)
            except ValueError:
                s = requests.Session()
                page = s.get("http://nseindia.com",headers=headers)
                page = s.get(choose_url(choice),headers=headers).json()

            if len(page)>0:
                break


        now = datetime.now()
        time = now.strftime("%H:%M:%S")

        print("Record Fetch Time =", time)        

        res = page['records']['data'] #########
        underlying_value = int(res[0]['PE']['underlyingValue']) #######

        if choice not in ['NIFTY','BANKNIFTY']:
            selected_margin = underlying_value * selected_margin / 100

        data = [d.get('CE') for d in res]
        data_cleaned = list(filter(None, data))
        df_call = pd.DataFrame(data_cleaned)

        data = [d.get('PE') for d in res]
        data_cleaned = list(filter(None, data))
        df_put = pd.DataFrame(data_cleaned)
        
        # import pdb;pdb.set_trace()

        df_put['expiryDate'] = pd.to_datetime(df_put['expiryDate'])
        df_call['expiryDate'] = pd.to_datetime(df_call['expiryDate'])

        df_put.sort_values(by=['expiryDate','strikePrice'],inplace=True)
        df_call.sort_values(by=['expiryDate','strikePrice'],inplace=True)
        
        col_list = ['strikePrice','expiryDate','openInterest']
        df_put = df_put[col_list]
        df_call = df_call[col_list]

        df_put = df_put[df_put['expiryDate'] == date]
        df_call = df_call[df_call['expiryDate'] == date]

        df_put['strikePrice'] = df_put['strikePrice'].map(int)
        df_call['strikePrice'] = df_call['strikePrice'].map(int)

        df_put = df_put[(df_put['strikePrice'] >= underlying_value-selected_margin) & (df_put['strikePrice']<=underlying_value+selected_margin)]
        df_call = df_call[(df_call['strikePrice'] >= underlying_value-selected_margin) & (df_call['strikePrice']<=underlying_value+selected_margin)]
        df_put['RIGHT'] = 'PUT'
        df_call['RIGHT'] = 'CALL'
        df_merged = df_put.merge(df_call, on = 'strikePrice')
        df_merged.insert(0, 'Time', time)

        cols = ['strikePrice','Time','Stock/Index','expiryDate_x','RIGHT_x','openInterest_x','previous_min_OI_x','change_OI_x',
                    '% Change_OI_x','RIGHT_y','openInterest_y','previous_min_OI_y','change_OI_y',
                    '% Change_OI_y']
    
        df_merged['Stock/Index'] = res[0]['PE']['underlying']
        df_merged['previous_min_OI_x'] = previous_min_OI_x
        df_merged['previous_min_OI_y'] = previous_min_OI_y
        df_merged['change_OI_x'] = df_merged['openInterest_x'] - df_merged['previous_min_OI_x']
        df_merged['change_OI_y'] =  df_merged['openInterest_y'] - df_merged['previous_min_OI_y']
        df_merged['% Change_OI_x'] = df_merged['change_OI_x']/df_merged['previous_min_OI_x'] * 100.0
        df_merged['% Change_OI_y'] = df_merged['change_OI_y']/df_merged['previous_min_OI_y'] * 100.0
        
        previous_min_OI_x = df_merged['openInterest_x'].values.tolist()
        previous_min_OI_y = df_merged['openInterest_y'].values.tolist()
            
        df_merged = df_merged[cols]
        file_name = choice + ' ' + date

        for i in sp_range:
            grouped = df_merged.groupby(df_merged.strikePrice)
            df_grouped = grouped.get_group(i)
            # df_grouped = df_grouped.style.format({'change_OI_x':"{0:+g}",'change_OI_y':"{0:+g}",'% Change_OI_x':"{0:+g}",'% Change_OI_y':"{0:+g}"})
            del df_grouped['strikePrice']        
            append_df_to_excel(f"{file_name}.xlsx",df_grouped,sheet_name=str(i), header=None, index=False)
    
    except:
        print('Exception response stored in json : ')
        # with open('res.json', 'w') as json_file:
        #     json.dump(res, json_file)

    return previous_min_OI_x,previous_min_OI_y
    ########################### MAIN CODE ##########################


choice = input('Enter Symbol / Index : ').upper()
margin = int(input('Enter Margin Value ( if STOCK give %)  : '))
date = input("Enter date in format yyyy-mm-dd : ")

sp_range = init_spreadsheet(choice,date,margin)

previous_min_OI_x = [0 for i in sp_range]
previous_min_OI_y = [0 for i in sp_range]

while(True):
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    print(current_time)
    previous_min_OI_x , previous_min_OI_y = main(choice,previous_min_OI_x,previous_min_OI_y,margin)
    sleep(83 - time() % 60)
